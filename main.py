import requests
from bs4 import BeautifulSoup
import json
import time
from openpyxl import Workbook
import pymysql
import uuid
import sys

# 106.12.106.197
# sto
# Msd190313
# stopricedata

basecustomerinfo = ""  # 核心ID，决定了数据存放归属于哪家站点的报价，不能弄错，每次需要修改
conn = pymysql.connect(host='', user='', passwd='', port=3306, db='', charset='utf8')
cur = conn.cursor()  # 配置好数据库访问，用于数据抓取后存放


# def read_datafile(filename):
#     with open(filename,"r",encoding='UTF-8') as f:
#         data = f.read()
#         print(data)
#     print(f'Hello world~')

# 读取网页数据
# 此处，针对于STO需要进行cookie的抓取和配置，否则无法登录进行数据抓取
def read_webdata(webadd):
    cookie = "cookie: cna=AvAmF6acrXwCAbcPskX0TjJ2; WD_SESSION=1c5e8714-ff5f-4256-ad46-47902e779e9a; " \
             "isg=BO_vsh4F7FpY6ehtQLx9g0oGfgP5lEO2a2TsrwF8g95lUA9SCWVvBrYS0EjuMxsu "

    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/75.0.3770.142 Safari/537.36',
        'Connection': 'keep-alive',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,'
                  'application/signed-exchange;v=b3',
        'Cookie': cookie,
        'referer': 'https://wangdian.sto.cn/api'
    }
    url = webadd  # 加载抓取信息的网页接口地址
    seesion = requests.session()
    response = seesion.get(url, headers=header)
    wbdata = response.text
    print("正在打开请求")
    print(response.url)
    soup = BeautifulSoup(wbdata, 'lxml')

    return soup


def insertOne(value, sheet):
    row = [basecustomerinfo, value['id'], value['quoteName'], value['quoteCode'],
           "https://wangdian.sto.cn/api/amount/transfer/get/stepthree?id=" + value['id'] + "&action=view",
           value['siteCode'], value['siteName'], value['usageState'], value['auditState'], value['priority'],
           value['quoteCategory'], value['transportString'], value['billTypeString'], value['billCategoryString'],
           value['goodsTypeString'], value['createBy'], value['createUserId'], value['modifiedBy'],
           value['modifiedUserId']]
    # print(type(row))   row是list类型
    sheet.append(row)
    cur.execute(
        "insert into zd_base_info (cusnum, priceid, pricename, pricenum, pricewebadd, parcusnum, parcusname, "
        "usestatus, singstatus, sortnum, pricetype, trntype, ordertype, ordmodel, protype, cc, ccnum, mc, "
        "mcnum) values('" + basecustomerinfo + "', '" + str(
            value['id']) + "', '" + str(value['quoteName']) + "', '" + str(
            value['quoteCode']) + "', '" + "https://wangdian.sto.cn/api/amount/transfer/get/stepthree?id=" + str(
            value['id']) + "&action=view" + "', '" + str(value['siteCode']) + "', '" + str(
            value['siteName']) + "', '" + str(value['usageState']) + "', '" + str(value['auditState']) + "', '" + str(
            value['priority']) + "', '" + str(value['quoteCategory']) + "', '" + str(
            value['transportString']) + "', '" + str(value['billTypeString']) + "', '" + str(
            value['billCategoryString']) + "', '" + str(value['goodsTypeString']) + "', '" + str(
            value['createBy']) + "', '" + str(value['createUserId']) + "', '" + str(value['modifiedBy']) + "', '" + str(
            value['modifiedUserId']) + "')")
    conn.commit()


def create_exceldoc(diclist):
    book = Workbook()

    sheet = book.create_sheet("申通报价资料", 0)
    sheet.append(
        ["所属区域客户编号", "报价ID", "报价名称", "报价编号", "报价网址", "所属站点编号", "所属站点名称", "使用状态", "审核状态", "优先级", "报价类型", "运输方式", "面单类型",
         "面单类别", "物品类别", "创建人", "创建人编号", "修改人", "修改人编号"])

    sheets = book.get_sheet_names()
    count = 0
    # 向sheet中插入数据
    # webaddrow=[]
    for its in diclist:
        count = count + 1
        insertOne(its, book.get_sheet_by_name(sheets[0]))

    # 保存数据到.xlsx文件
    # book.save("d:\\test.xlsx")
    print(str(count))
    return diclist


def create_priceexceldoc(diclist, fileid, book):
    count = 0

    for x in diclist:
        for y in x["area"]:
            row = [fileid, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
            for yy in y["citys"]:
                row[1] = y["provinceName"]
                row[2] = y["provinceId"]
                row[3] = yy["cityName"]
                row[4] = yy["cityId"]
                for z in x["weight"]:
                    row[5] = str(z["startWeight"]) + "<=w<=" + str(z["endWeight"])
                    row[6] = z["priceExpression"]
                    row[7] = z["weightModeTypeG"]
                    row[8] = z["piecePrice"]
                    row[9] = z["continuedHeavy"]
                    row[10] = z["continuedHeavyPrice"]
                    row[11] = z["surcharge"]
                    row[12] = z["lowestPrice"]
                    row[13] = z["weightModeNameG"]
                    row[14] = z["weightCarriesNumberG"]
                    row[15] = z["weightDiscardsNumberG"]
                    row[16] = z["weightModeParameterG"]
                    row[17] = z["ykg"]
                    row[18] = z["ykgPrice"]
                    book.get_sheet_by_name(sheets[0]).append(row)
                    cur.execute(
                        "INSERT INTO pricelist (id, priceid, provname, provid, cityname, cityid, weirog, pricemodel, "
                        "wetmodel, piecePrice, continuedHeavy, continuedHeavyPrice, surcharge, lowestPrice, "
                        "weightModeNameG, weightCarriesNumberG, weightDiscardsNumberG, weightModeParameterG, ykg, "
                        "ykgPrice) VALUES ('" + str(
                            uuid.uuid4()) + "', '" + str(fileid) + "', '" + str(row[1]) + "', '" + str(
                            row[2]) + "', '" + str(row[3]) + "', '" + str(row[4]) + "', '" + str(row[5]) + "', '" + str(
                            row[6]) + "', '" + str(row[7]) + "', '" + str(row[8]) + "', '" + str(row[9]) + "', '" + str(
                            row[10]) + "', '" + str(row[11]) + "', '" + str(row[12]) + "', '" + str(
                            row[13]) + "', '" + str(row[14]) + "', '" + str(row[15]) + "', '" + str(
                            row[16]) + "', '" + str(row[17]) + "', '" + str(row[18]) + "')")
                    conn.commit()
                    count = count + 1
    return book


if __name__ == '__main__':
    if len(basecustomerinfo) < 1:
        print("程序没有完成初始化配置，退出程序")
        sys.exit(0)
    webaddrow = {}  # 关于明细报价表格的行数据集基础容器

    # 获取当前爬取数据基础信息
    jsonbase = json.loads(read_webdata(
        'https://wangdian.sto.cn/api/amount/transfer/searchQuotePrice?current=999&pageSize=100&params=%7B%22siteCode'
        '%22%3A%22471000%22%2C%22feeType%22%3A%220%22%7D').text)
    totalinfo = jsonbase["data"]["paging"]["total"]
    totalpage = jsonbase["data"]["paging"]["pageCount"]

    # 根据基础信息爬取所有页面，站点报价基础信息
    alllist = []
    for num in range(1, totalpage + 1):
        jsonobj = json.loads(read_webdata('https://wangdian.sto.cn/api/amount/transfer/searchQuotePrice?current=' + str(
            num) + '&pageSize=100&params=%7B%22siteCode%22%3A%22471000%22%2C%22feeType%22%3A%220%22%7D').text)
        # print(type(jsonobj))
        jsonobj.pop('success')
        jsonobj.pop('requestId')
        if len(jsonobj.items()) == 1:
            for ite in jsonobj.items():
                # print(type(ite[1]['items']))
                alllist.extend(ite[1]['items'])
        time.sleep(3)
    webaddrow = create_exceldoc(alllist)

    # 创建excel文件接口
    wb = Workbook()  # book = Workbook()
    # 声明excel文件簿表名，及表头
    sheet = wb.create_sheet("申通报价明细资料", 0)
    sheet.append(
        ["报价ID", "省", "省编号", "市", "市编号", "重量区间", "价格公式", "重量模式", "piecePrice", "continuedHeavy", "continuedHeavyPrice",
         "surcharge", "lowestPrice", "weightModeNameG", "weightCarriesNumberG", "weightDiscardsNumberG",
         "weightModeParameterG", "ykg", "ykgPrice"])
    # 循环站点列表，通过ID，获取相关站点报价明细信息
    sheets = wb.get_sheet_names()
    for p in webaddrow:
        jsonobjc = json.loads(read_webdata(
            'https://wangdian.sto.cn/api/amount/transfer/get/stepthree?id=' + p['id'] + '&action=view').text)
        wb = create_priceexceldoc(jsonobjc['data']['stoQuoteRegion'], p['id'], wb)
        time.sleep(3)  # 每次循环间歇3秒

    # 保存数据到.xlsx文件
    # wb.save("d:\\.xlsx")

    conn.close()
